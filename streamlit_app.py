import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from io import BytesIO
import io
import json
from pathlib import Path

st.set_page_config(page_title="CTR Processor", layout="wide")

# Load company indicators from config file
@st.cache_resource
def load_company_config():
    """Load company indicators from company_config.json"""
    config_path = Path(__file__).parent / "company_config.json"
    try:
        with open(config_path, 'r') as f:
            config = json.load(f)
        return config.get('companies', {})
    except FileNotFoundError:
        st.warning(f"Config file not found at {config_path}. Using default companies.")
        return {
            'Bank of America': ['bank of america', 'bofa', 'b of a'],
            'Wells Fargo': ['wells fargo', 'wellsfargo'],
            'Citibank': ['citibank', 'citi'],
            'Chase': ['chase', 'jp morgan chase', 'jpmorgan'],
            'Capital One': ['capital one'],
            'American Express': ['american express', 'amex'],
        }

def extract_company_name_from_keywords(keywords_list):
    """Fallback: Extract company name from keywords using capitalization."""
    keywords_str = [str(kw).strip() for kw in keywords_list if pd.notna(kw)]
    potential_companies = []
    for kw in keywords_str:
        if kw and kw[0].isupper():
            potential_companies.append(kw)
    
    if not potential_companies:
        return None
    
    from collections import Counter
    company_freq = Counter(potential_companies)
    
    if company_freq:
        return company_freq.most_common(1)[0][0]
    return None

def auto_learn_company(company_name, keywords_list):
    """Automatically add new company to config."""
    keywords_lower = [str(kw).lower().strip() for kw in keywords_list if pd.notna(kw)]
    
    company_indicators = load_company_config()
    
    if company_name in company_indicators:
        return False
    
    from collections import Counter
    keyword_freq = Counter(keywords_lower)
    
    learned_keywords = [kw for kw, count in keyword_freq.most_common(10) if kw and len(kw) > 2]
    
    if learned_keywords:
        company_indicators[company_name] = learned_keywords
        save_company_config(company_indicators)
        return True
    
    return False

def save_company_config(company_indicators):
    """Save updated company config back to JSON file."""
    config_path = Path(__file__).parent / "company_config.json"
    config = {"companies": company_indicators}
    with open(config_path, 'w') as f:
        json.dump(config, f, indent=2)
    
    st.cache_resource.clear()



def identify_company(keywords_list):
    """
    Identify company from keywords using configuration and pattern matching.
    """
    keywords_lower = [str(kw).lower().strip() for kw in keywords_list if pd.notna(kw)]
    config_companies = load_company_config()
    
    # Check keywords against known company indicators
    keyword_company_matches = {}
    for keyword in keywords_lower:
        for company, indicators in config_companies.items():
            if keyword in indicators:
                keyword_company_matches[company] = keyword_company_matches.get(company, 0) + 1
    
    # Return the most matched company
    if keyword_company_matches:
        return max(keyword_company_matches, key=keyword_company_matches.get)
    
    # Fallback to extracting from keywords
    return extract_company_name_from_keywords(keywords_list)

def process_excel_files(uploaded_files):
    """
    Process all uploaded Excel files.
    Returns aggregated data by company.
    """
    if not uploaded_files:
        return None
    
    monthly_data = []
    company_aggregates = {}
    newly_detected = []
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    for idx, uploaded_file in enumerate(uploaded_files):
        status_text.text(f"Processing: {uploaded_file.name}")
        
        try:
            df = pd.read_excel(uploaded_file, sheet_name=0)
            
            # Extract columns D (index 3) and H (index 7)
            search_volume_col = df.iloc[:, 3]  # Column D
            traffic_col = df.iloc[:, 7]  # Column H
            
            # Get keywords from first column to identify company
            keywords = df.iloc[:, 0]
            
            # Identify company from keywords
            company = identify_company(keywords.tolist())
            
            # Track if this is a newly detected company
            config_companies = load_company_config()
            if company not in config_companies and company != 'Unknown Company':
                if company not in newly_detected:
                    newly_detected.append(company)
            
            # Calculate monthly totals
            monthly_search_volume = pd.to_numeric(search_volume_col, errors='coerce').sum()
            monthly_traffic = pd.to_numeric(traffic_col, errors='coerce').sum()
            
            # Store monthly data
            monthly_data.append({
                'File Name': uploaded_file.name,
                'Company': company,
                'Monthly Search Volume': int(monthly_search_volume) if pd.notna(monthly_search_volume) else 0,
                'Monthly Traffic': int(monthly_traffic) if pd.notna(monthly_traffic) else 0
            })
            
            # Aggregate by company for yearly calculations
            if company not in company_aggregates:
                company_aggregates[company] = {
                    'Total Search Volume': 0,
                    'Total Traffic': 0
                }
            
            company_aggregates[company]['Total Search Volume'] += monthly_search_volume
            company_aggregates[company]['Total Traffic'] += monthly_traffic
            
        except Exception as e:
            st.error(f"Error processing {uploaded_file.name}: {str(e)}")
            continue
        
        progress_bar.progress((idx + 1) / len(uploaded_files))
    
    status_text.empty()
    progress_bar.empty()
    
    # Calculate CTR and prepare company summary
    company_summary = []
    for company, data in sorted(company_aggregates.items()):
        total_search_volume = data['Total Search Volume']
        total_traffic = data['Total Traffic']
        ctr = total_traffic / total_search_volume if total_search_volume > 0 else 0
        
        company_summary.append({
            'Company': company,
            'Total Search Volume': int(total_search_volume),
            'Total Traffic': int(total_traffic),
            'CTR': ctr
        })
    
    return {
        'company_summary': company_summary,
        'monthly_data': monthly_data,
        'newly_detected': newly_detected
    }

def generate_output_excel(results):
    """
    Generate the output Excel file with two sheets.
    Returns BytesIO object.
    """
    if not results:
        return None
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Sheet 1: Company Summary
    ws1 = wb.create_sheet('Company Summary', 0)
    ws1.append(['Company', 'Total Search Volume', 'Total Traffic', 'CTR'])
    
    for row in results['company_summary']:
        ws1.append([
            row['Company'],
            row['Total Search Volume'],
            row['Total Traffic'],
            row['CTR']
        ])
    
    # Sheet 2: Monthly Summary
    ws2 = wb.create_sheet('Monthly Summary', 1)
    ws2.append(['File Name', 'Company', 'Monthly Search Volume', 'Monthly Traffic'])
    
    for row in results['monthly_data']:
        ws2.append([
            row['File Name'],
            row['Company'],
            row['Monthly Search Volume'],
            row['Monthly Traffic']
        ])
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Streamlit UI
st.title("📊 CTR Processor")
st.markdown("Upload multiple Excel files to calculate yearly CTR by company")

st.divider()

# Sidebar: Information
st.sidebar.subheader("ℹ️ How It Works")
st.sidebar.markdown("""
This app:
1. Reads Excel files with keywords
2. Matches keywords to companies
3. Calculates CTR by company
""")

# File upload section
st.subheader("Step 1: Upload Excel Files")
uploaded_files = st.file_uploader(
    "Choose Excel files (.xlsx, .xls)",
    type=['xlsx', 'xls'],
    accept_multiple_files=True
)

if uploaded_files:
    st.success(f"✅ {len(uploaded_files)} file(s) uploaded")
    
    # Process button
    if st.button("🔄 Process Files", type="primary", use_container_width=True):
        st.divider()
        results = process_excel_files(uploaded_files)
            
        if results:
            st.session_state['results'] = results
            st.success("✅ Files processed successfully!")
            
            # Display results
            st.subheader("Company Summary")
            summary_df = pd.DataFrame(results['company_summary'])
            st.dataframe(summary_df, use_container_width=True, hide_index=True)
            
            st.subheader("Monthly Summary")
            monthly_df = pd.DataFrame(results['monthly_data'])
            st.dataframe(monthly_df, use_container_width=True, hide_index=True)
            
            st.divider()
            
            # Download button
            excel_file = generate_output_excel(results)
            st.download_button(
                label="📥 Download CTR_Summary.xlsx",
                data=excel_file,
                file_name="CTR_Summary.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
else:
    st.info("👆 Upload Excel files to get started")

st.divider()

# Footer with information
st.markdown("""
---
**📊 How CTR is Calculated:**
- Upload Excel files with company keywords and data
- Keywords are matched against company configurations
- Companies are identified and data is aggregated
- New companies can be added to company_config.json
- CTR is calculated as: Traffic ÷ Search Volume
""")
