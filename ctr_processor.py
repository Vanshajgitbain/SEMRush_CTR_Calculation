import pandas as pd
import openpyxl
from openpyxl.styles import Font
from pathlib import Path
import os

def identify_company(keywords_list):
    """
    Identify company from keywords in the file.
    Returns the identified company name.
    """
    # Convert all keywords to lowercase for matching
    keywords_lower = [str(kw).lower() for kw in keywords_list if pd.notna(kw)]
    
    company_indicators = {
        'Bank of America': ['bank of america', 'bofa', 'b of a'],
        'Wells Fargo': ['wells fargo', 'wellsfargo'],
        'Citibank': ['citibank', 'citi'],
        'Chase': ['chase', 'jp morgan chase', 'jpmorgan'],
        'Capital One': ['capital one'],
        'American Express': ['american express', 'amex'],
        # Add more companies as needed
    }
    
    company_counts = {company: 0 for company in company_indicators}
    
    for keyword in keywords_lower:
        for company, indicators in company_indicators.items():
            for indicator in indicators:
                if indicator in keyword:
                    company_counts[company] += 1
    
    # Return company with highest match count
    if max(company_counts.values()) > 0:
        return max(company_counts, key=company_counts.get)
    else:
        return 'Unknown Company'

def process_excel_files(directory_path):
    """
    Process all Excel files in the specified directory.
    Returns aggregated data by company.
    """
    excel_files = list(Path(directory_path).glob('*.xlsx')) + list(Path(directory_path).glob('*.xls'))
    
    if not excel_files:
        print(f"No Excel files found in {directory_path}")
        return None
    
    print(f"Found {len(excel_files)} Excel files to process")
    
    monthly_data = []  # For Sheet 2
    company_aggregates = {}  # For Sheet 1
    
    for file_path in sorted(excel_files):
        print(f"Processing: {file_path.name}")
        
        try:
            df = pd.read_excel(file_path, sheet_name=0)
            
            # Extract columns D (index 3) and H (index 7)
            search_volume_col = df.iloc[:, 3]  # Column D
            traffic_col = df.iloc[:, 7]  # Column H
            
            # Get keywords from first column to identify company
            keywords = df.iloc[:, 0]  # Assuming first column has keywords
            
            # Identify company
            company = identify_company(keywords.tolist())
            
            # Calculate monthly totals
            monthly_search_volume = pd.to_numeric(search_volume_col, errors='coerce').sum()
            monthly_traffic = pd.to_numeric(traffic_col, errors='coerce').sum()
            
            # Store monthly data
            monthly_data.append({
                'File Name': file_path.name,
                'Company': company,
                'Monthly Search Volume': monthly_search_volume,
                'Monthly Traffic': monthly_traffic
            })
            
            # Aggregate by company for yearly calculations
            if company not in company_aggregates:
                company_aggregates[company] = {
                    'Total Search Volume': 0,
                    'Total Traffic': 0
                }
            
            company_aggregates[company]['Total Search Volume'] += monthly_search_volume
            company_aggregates[company]['Total Traffic'] += monthly_traffic
            
            print(f"  Company: {company}")
            print(f"  Monthly Search Volume: {monthly_search_volume}")
            print(f"  Monthly Traffic: {monthly_traffic}")
            
        except Exception as e:
            print(f"  Error processing {file_path.name}: {str(e)}")
            continue
    
    # Calculate CTR and prepare company summary
    company_summary = []
    for company, data in sorted(company_aggregates.items()):
        total_search_volume = data['Total Search Volume']
        total_traffic = data['Total Traffic']
        ctr = total_traffic / total_search_volume if total_search_volume > 0 else 0
        
        company_summary.append({
            'Company': company,
            'Total Search Volume': total_search_volume,
            'Total Traffic': total_traffic,
            'CTR': ctr
        })
    
    return {
        'company_summary': company_summary,
        'monthly_data': monthly_data
    }

def generate_output_excel(results, output_path):
    """
    Generate the output Excel file with two sheets:
    Sheet 1: Company Summary
    Sheet 2: Monthly Summary
    """
    if not results:
        print("No data to export")
        return
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Sheet 1: Company Summary
    ws1 = wb.create_sheet('Company Summary', 0)
    ws1.append(['Company', 'Total Search Volume', 'Total Traffic', 'CTR'])
    
    for row in results['company_summary']:
        ws1.append([
            row['Company'],
            row['Total Search Volume'],
            row['Total Traffic'],
            row['CTR']
        ])
    
    # Sheet 2: Monthly Summary
    ws2 = wb.create_sheet('Monthly Summary', 1)
    ws2.append(['File Name', 'Company', 'Monthly Search Volume', 'Monthly Traffic'])
    
    for row in results['monthly_data']:
        ws2.append([
            row['File Name'],
            row['Company'],
            row['Monthly Search Volume'],
            row['Monthly Traffic']
        ])
    
    # Save workbook
    wb.save(output_path)
    print(f"\nOutput file generated: {output_path}")

def main():
    # Set the directory where Excel files are located
    input_directory = r"c:\Users\78594\OneDrive - Bain\Documents\Training\New folder"
    output_file = r"c:\Users\78594\OneDrive - Bain\Documents\Training\New folder\CTR_Summary.xlsx"
    
    print("=" * 60)
    print("CTR Processor - Multiple Excel Files")
    print("=" * 60)
    
    # Process files
    results = process_excel_files(input_directory)
    
    if results:
        print("\n" + "=" * 60)
        print("AGGREGATION RESULTS")
        print("=" * 60)
        
        print("\nCompany Summary:")
        for row in results['company_summary']:
            print(f"  {row['Company']}: Search Volume={row['Total Search Volume']}, Traffic={row['Total Traffic']}, CTR={row['CTR']:.4f}")
        
        print(f"\nTotal Monthly Records: {len(results['monthly_data'])}")
        
        # Generate output Excel
        generate_output_excel(results, output_file)
        print("\nProcessing complete!")
    else:
        print("No files were processed.")

if __name__ == "__main__":
    main()
